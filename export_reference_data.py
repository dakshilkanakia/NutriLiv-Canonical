#!/usr/bin/env python3
"""
Export Reference Data to JSON for Flutter App

Converts Excel reference tables to optimized JSON files for bundling in Flutter app.

Usage:
    python export_reference_data.py

Outputs:
    - flutter_assets/ingredients.json
    - flutter_assets/display_policies.json
    - flutter_assets/densities.json
    - flutter_assets/conversion_constants.json
"""

import openpyxl
import json
import os
from typing import Dict, List
from config import MASS_TO_GRAMS, VOLUME_TO_ML, UNIT_DIMENSIONS, INGREDIENT_TABLE, DENSITY_TABLE

# Output directory
OUTPUT_DIR = "flutter_assets"

def create_output_dir():
    """Create output directory if it doesn't exist"""
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        print(f"✅ Created directory: {OUTPUT_DIR}/")
    else:
        print(f"📁 Using existing directory: {OUTPUT_DIR}/")


def export_ingredients():
    """Export ingredients table to JSON"""
    print("\n" + "=" * 80)
    print("EXPORTING INGREDIENTS")
    print("=" * 80)
    
    filepath = INGREDIENT_TABLE
    
    if not os.path.exists(filepath):
        print(f"❌ ERROR: File not found: {filepath}")
        return
    
    print(f"📖 Reading: {filepath}")
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    ingredients = {}
    
    # Read header to get column indices
    headers = [cell.value for cell in ws[1]]
    
    for row_idx in range(2, ws.max_row + 1):
        ingredient_id = ws.cell(row_idx, 1).value
        primary_name = ws.cell(row_idx, 2).value
        aliases_raw = ws.cell(row_idx, 3).value
        category = ws.cell(row_idx, 4).value
        default_form_id = ws.cell(row_idx, 5).value
        
        if not ingredient_id or not primary_name:
            continue
        
        # Parse aliases
        aliases = []
        if aliases_raw:
            aliases = [a.strip() for a in str(aliases_raw).split(";") if a.strip()]
        
        ingredients[ingredient_id] = {
            "ingredient_id": ingredient_id,     # Make ID available to Dart models
            "primary_name": primary_name,      # Match Dart model field name
            "category": category,
            "default_form": default_form_id,
            "aliases": aliases
        }
    
    # Write to JSON
    output_file = os.path.join(OUTPUT_DIR, "ingredients.json")
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(ingredients, f, indent=2, ensure_ascii=False)
    
    print(f"✅ Exported {len(ingredients)} ingredients")
    print(f"📝 Output: {output_file}")
    
    # Calculate file size
    file_size = os.path.getsize(output_file)
    print(f"📦 Size: {file_size / 1024:.1f} KB")


def export_display_policies():
    """Export display policies to JSON"""
    print("\n" + "=" * 80)
    print("EXPORTING DISPLAY POLICIES")
    print("=" * 80)
    
    filepath = "Ingredient Display Policy 251202.xlsx"
    
    if not os.path.exists(filepath):
        print(f"❌ ERROR: File not found: {filepath}")
        return
    
    print(f"📖 Reading: {filepath}")
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    policies = {}
    
    for row_idx in range(2, ws.max_row + 1):
        ingredient_id = ws.cell(row_idx, 1).value
        default_display = ws.cell(row_idx, 2).value
        locale_overrides_raw = ws.cell(row_idx, 3).value
        rationale = ws.cell(row_idx, 4).value
        
        if not ingredient_id:
            continue
        
        # Parse locale overrides from JSON string
        locale_overrides = {}
        if locale_overrides_raw and str(locale_overrides_raw).strip():
            try:
                # Handle the format: {"en-US":"prefer_volume", "en-CA":"prefer_volume"}
                locale_overrides = json.loads(locale_overrides_raw)
            except json.JSONDecodeError:
                # If not valid JSON, skip
                pass
        
        policies[ingredient_id] = {
            "default_display_rule": default_display if default_display else "prefer_mass",  # Match Dart model
            "locale_overrides": locale_overrides,
            "rationale": rationale if rationale else ""
        }
    
    # Write to JSON
    output_file = os.path.join(OUTPUT_DIR, "display_policies.json")
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(policies, f, indent=2, ensure_ascii=False)
    
    print(f"✅ Exported {len(policies)} display policies")
    print(f"📝 Output: {output_file}")
    
    # Calculate file size
    file_size = os.path.getsize(output_file)
    print(f"📦 Size: {file_size / 1024:.1f} KB")


def export_densities():
    """Export density table to JSON"""
    print("\n" + "=" * 80)
    print("EXPORTING DENSITIES")
    print("=" * 80)
    
    filepath = DENSITY_TABLE
    
    if not os.path.exists(filepath):
        print(f"❌ ERROR: File not found: {filepath}")
        return
    
    print(f"📖 Reading: {filepath}")
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    densities = {}
    
    for row_idx in range(2, ws.max_row + 1):
        density_id = ws.cell(row_idx, 1).value
        ingredient_id = ws.cell(row_idx, 2).value
        form_id = ws.cell(row_idx, 3).value
        g_per_ml = ws.cell(row_idx, 4).value
        
        if not density_id or not ingredient_id or not form_id or g_per_ml is None:
            continue
        
        # Use composite key: ingredient_id|form_id (match Dart lookup key)
        key = f"{ingredient_id}|{form_id}"
        
        densities[key] = {
            "density_id": density_id,
            "ingredient_id": ingredient_id,
            "form_id": form_id,
            "density_g_per_ml": float(g_per_ml)  # Match Dart model field name
        }
    
    # Write to JSON
    output_file = os.path.join(OUTPUT_DIR, "densities.json")
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(densities, f, indent=2, ensure_ascii=False)
    
    print(f"✅ Exported {len(densities)} density entries")
    print(f"📝 Output: {output_file}")
    
    # Calculate file size
    file_size = os.path.getsize(output_file)
    print(f"📦 Size: {file_size / 1024:.1f} KB")


def export_conversion_constants():
    """Export unit conversion constants to JSON"""
    print("\n" + "=" * 80)
    print("EXPORTING CONVERSION CONSTANTS")
    print("=" * 80)
    
    constants = {
        "mass_to_grams": MASS_TO_GRAMS,
        "volume_to_ml": VOLUME_TO_ML,
        "unit_dimensions": UNIT_DIMENSIONS,
        "version": "1.0",
        "last_updated": "2024-12-01"
    }
    
    # Write to JSON
    output_file = os.path.join(OUTPUT_DIR, "conversion_constants.json")
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(constants, f, indent=2, ensure_ascii=False)
    
    print(f"✅ Exported conversion constants")
    print(f"   - Mass conversions: {len(MASS_TO_GRAMS)} units")
    print(f"   - Volume conversions: {len(VOLUME_TO_ML)} units")
    print(f"   - Unit dimensions: {len(UNIT_DIMENSIONS)} units")
    print(f"📝 Output: {output_file}")
    
    # Calculate file size
    file_size = os.path.getsize(output_file)
    print(f"📦 Size: {file_size / 1024:.1f} KB")


def print_summary():
    """Print summary of exported files"""
    print("\n" + "=" * 80)
    print("EXPORT COMPLETE!")
    print("=" * 80)
    
    if not os.path.exists(OUTPUT_DIR):
        print("❌ No files exported")
        return
    
    files = [
        "ingredients.json",
        "display_policies.json",
        "densities.json",
        "conversion_constants.json"
    ]
    
    total_size = 0
    print("\n📦 Exported files:")
    for filename in files:
        filepath = os.path.join(OUTPUT_DIR, filename)
        if os.path.exists(filepath):
            size = os.path.getsize(filepath)
            total_size += size
            print(f"   ✅ {filename:30} ({size / 1024:.1f} KB)")
        else:
            print(f"   ❌ {filename:30} (not found)")
    
    print(f"\n📊 Total size: {total_size / 1024:.1f} KB")
    print(f"📁 Location: {OUTPUT_DIR}/")
    
    print("\n🎯 Next steps:")
    print("   1. Review the JSON files in flutter_assets/")
    print("   2. Copy flutter_assets/ to your Flutter project's assets/")
    print("   3. Add to pubspec.yaml:")
    print("      assets:")
    print("        - assets/flutter_assets/")
    print("\n" + "=" * 80)


def main():
    """Main export function"""
    print("\n" + "=" * 80)
    print("REFERENCE DATA EXPORT TOOL")
    print("Converting Excel files → JSON for Flutter App")
    print("=" * 80)
    
    # Create output directory
    create_output_dir()
    
    # Export each data source
    export_ingredients()
    export_display_policies()
    export_densities()
    export_conversion_constants()
    
    # Print summary
    print_summary()


if __name__ == "__main__":
    main()

