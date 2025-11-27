"""
Stage 2 Processor - Data Loaders
Loads reference data from Excel files
"""

import openpyxl
import json
from typing import Dict, List, Tuple
import os

class ReferenceData:
    """Container for all reference data"""
    def __init__(self):
        self.ingredients = {}  # ingredient_id -> dict
        self.ingredients_by_name = {}  # normalized_name -> ingredient_id
        self.ingredients_aliases = {}  # alias -> ingredient_id
        
        self.forms = {}  # form_id -> dict
        self.densities = {}  # (ingredient_id, form_id) -> density_data
        self.conversions = {}  # (from_unit, to_unit) -> factor
        
        self.meaning_tokens = {}  # category -> list of tokens


def load_ingredient_table(filepath: str) -> Tuple[Dict, Dict, Dict]:
    """
    Load ingredient table from Excel
    Returns: (ingredients_dict, name_index, alias_index)
    """
    print(f"Loading ingredient table from: {filepath}")
    
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Ingredient table not found: {filepath}")
    
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    ingredients = {}
    name_index = {}
    alias_index = {}
    
    # Skip header row
    for row_idx in range(2, ws.max_row + 1):
        ingredient_id = ws.cell(row_idx, 1).value
        primary_name = ws.cell(row_idx, 2).value
        aliases_raw = ws.cell(row_idx, 3).value
        category = ws.cell(row_idx, 4).value
        default_form_id = ws.cell(row_idx, 5).value
        
        if not ingredient_id or not primary_name:
            continue
        
        # Store ingredient data
        ingredients[ingredient_id] = {
            "ingredient_id": ingredient_id,
            "primary_name": primary_name,
            "category": category,
            "default_form_id": default_form_id,
            "aliases": []
        }
        
        # Index by normalized name
        normalized_name = normalize_for_matching(primary_name)
        name_index[normalized_name] = ingredient_id
        
        # Parse and index aliases
        if aliases_raw:
            aliases = [a.strip() for a in str(aliases_raw).split(";") if a.strip()]
            ingredients[ingredient_id]["aliases"] = aliases
            
            for alias in aliases:
                normalized_alias = normalize_for_matching(alias)
                alias_index[normalized_alias] = ingredient_id
    
    print(f"  Loaded {len(ingredients)} ingredients")
    return ingredients, name_index, alias_index


def load_form_table(filepath: str) -> Dict:
    """Load form table from Excel"""
    print(f"Loading form table from: {filepath}")
    
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Form table not found: {filepath}")
    
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    forms = {}
    
    for row_idx in range(2, ws.max_row + 1):
        form_id = ws.cell(row_idx, 1).value
        label = ws.cell(row_idx, 2).value
        description = ws.cell(row_idx, 3).value
        affects_density = ws.cell(row_idx, 4).value
        target_dimension = ws.cell(row_idx, 5).value
        
        if not form_id:
            continue
        
        forms[form_id] = {
            "form_id": form_id,
            "label": label,
            "description": description,
            "affects_density": affects_density,
            "target_dimension": target_dimension
        }
    
    print(f"  Loaded {len(forms)} forms")
    return forms


def load_density_table(filepath: str) -> Dict:
    """
    Load density table from Excel
    Returns: dict with (ingredient_id, form_id) -> density_data
    """
    print(f"Loading density table from: {filepath}")
    
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Density table not found: {filepath}")
    
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    densities = {}
    
    for row_idx in range(2, ws.max_row + 1):
        density_id = ws.cell(row_idx, 1).value
        ingredient_id = ws.cell(row_idx, 2).value
        form_id = ws.cell(row_idx, 3).value
        g_per_ml = ws.cell(row_idx, 4).value
        notes = ws.cell(row_idx, 5).value
        source = ws.cell(row_idx, 6).value
        
        if not density_id or not ingredient_id or not form_id or g_per_ml is None:
            continue
        
        key = (ingredient_id, form_id)
        densities[key] = {
            "density_id": density_id,
            "ingredient_id": ingredient_id,
            "form_id": form_id,
            "g_per_ml": float(g_per_ml),
            "notes": notes,
            "source": source
        }
    
    print(f"  Loaded {len(densities)} density entries")
    return densities


def load_conversion_table(filepath: str) -> Dict:
    """Load unit conversion constants"""
    print(f"Loading conversion table from: {filepath}")
    
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Conversion table not found: {filepath}")
    
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    conversions = {}
    
    for row_idx in range(2, ws.max_row + 1):
        from_unit = ws.cell(row_idx, 1).value
        to_unit = ws.cell(row_idx, 2).value
        factor = ws.cell(row_idx, 3).value
        
        if not from_unit or not to_unit or factor is None:
            continue
        
        key = (from_unit, to_unit)
        conversions[key] = float(factor)
    
    print(f"  Loaded {len(conversions)} conversion factors")
    return conversions


def load_meaning_tokens(filepath: str) -> Dict:
    """Load meaning-carrying tokens from JSON file"""
    print(f"Loading meaning tokens from: {filepath}")
    
    if not os.path.exists(filepath):
        print(f"  Warning: Meaning tokens file not found, using empty set")
        return {}
    
    with open(filepath, 'r', encoding='utf-8') as f:
        tokens = json.load(f)
    
    total = sum(len(v) for v in tokens.values())
    print(f"  Loaded {total} meaning tokens across {len(tokens)} categories")
    return tokens


def normalize_for_matching(text: str) -> str:
    """
    Normalize text for ingredient matching
    - Lowercase
    - Remove extra spaces
    - Basic singularization
    """
    if not text:
        return ""
    
    text = text.lower().strip()
    
    # Remove extra spaces
    text = " ".join(text.split())
    
    # Basic singularization (simple rules)
    if text.endswith("ies"):
        text = text[:-3] + "y"
    elif text.endswith("es") and len(text) > 3:
        text = text[:-2]
    elif text.endswith("s") and not text.endswith("ss"):
        text = text[:-1]
    
    return text


def load_all_reference_data(base_path: str = "/mnt/project") -> ReferenceData:
    """
    Load all reference data from Excel files
    
    Args:
        base_path: Directory containing the Excel files
    """
    print("\n" + "=" * 80)
    print("LOADING REFERENCE DATA")
    print("=" * 80)
    
    data = ReferenceData()
    
    # Import config to get filenames
    import sys
    sys.path.append(os.path.dirname(__file__))
    from config import (
        INGREDIENT_TABLE, FORM_TABLE, DENSITY_TABLE, 
        CONVERSION_TABLE, MEANING_TOKENS_FILE
    )
    
    # Load all tables
    try:
        ingredients, name_idx, alias_idx = load_ingredient_table(
            os.path.join(base_path, INGREDIENT_TABLE)
        )
        data.ingredients = ingredients
        data.ingredients_by_name = name_idx
        data.ingredients_aliases = alias_idx
        
        data.forms = load_form_table(os.path.join(base_path, FORM_TABLE))
        data.densities = load_density_table(os.path.join(base_path, DENSITY_TABLE))
        data.conversions = load_conversion_table(os.path.join(base_path, CONVERSION_TABLE))
        data.meaning_tokens = load_meaning_tokens(os.path.join(base_path, MEANING_TOKENS_FILE))
        
        print("\n" + "=" * 80)
        print("REFERENCE DATA LOADED SUCCESSFULLY")
        print("=" * 80)
        print(f"  Ingredients: {len(data.ingredients)}")
        print(f"  Forms: {len(data.forms)}")
        print(f"  Densities: {len(data.densities)}")
        print(f"  Conversions: {len(data.conversions)}")
        print("=" * 80 + "\n")
        
        return data
        
    except FileNotFoundError as e:
        print(f"\nERROR: {e}")
        print("Please ensure all reference files are in the correct location.")
        raise
    except Exception as e:
        print(f"\nERROR loading reference data: {e}")
        raise


if __name__ == "__main__":
    # Test loading
    data = load_all_reference_data()
    print("Test complete!")
