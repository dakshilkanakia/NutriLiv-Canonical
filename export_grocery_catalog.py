#!/usr/bin/env python3
"""
Export Grocery Item Catalog to JSON for Firebase

Converts Excel grocery catalog to JSON format for uploading to Firebase.

Usage:
    python export_grocery_catalog.py

Outputs:
    - flutter_assets/grocery_item_catalog.json
"""

import openpyxl
import json
import os
from typing import Dict, List, Any

# Output directory
OUTPUT_DIR = "flutter_assets"

def create_output_dir():
    """Create output directory if it doesn't exist"""
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        print(f"✅ Created directory: {OUTPUT_DIR}/")
    else:
        print(f"📁 Using existing directory: {OUTPUT_DIR}/")


def export_grocery_catalog():
    """Export grocery item catalog to JSON"""
    print("\n" + "=" * 80)
    print("EXPORTING GROCERY ITEM CATALOG")
    print("=" * 80)
    
    # Try Excel first, then CSV fallback
    excel_path = "Grocery Item Catalog 251226.xlsx"
    csv_path = "Grocery Item Catalog 251226.csv"
    
    # Try CSV first (more reliable)
    if os.path.exists(csv_path):
        print(f"📖 Reading CSV: {csv_path}")
        if export_from_csv(csv_path):
            return
    
    # Fallback to Excel
    if os.path.exists(excel_path):
        print(f"📖 Reading Excel: {excel_path}")
        if export_from_excel(excel_path):
            return
    
    print(f"❌ ERROR: Neither Excel nor CSV file found")
    print(f"   Looked for: {csv_path}")
    print(f"   Looked for: {excel_path}")
    print(f"\n💡 TIP: Please export your Excel file to CSV format:")
    print(f"   1. Open 'Grocery Item Catalog 251226.xlsx' in Excel")
    print(f"   2. File → Save As → CSV (Comma delimited)")
    print(f"   3. Save as 'Grocery Item Catalog 251226.csv' in the same directory")


def export_from_excel(filepath: str) -> bool:
    """Export from Excel file"""
    try:
        print("   📖 Loading workbook (read-only mode for large files)...")
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        
        # Try to get the first sheet
        if not wb.sheetnames:
            print("⚠️  No sheets found in Excel file, trying CSV...")
            return False
        
        ws = wb[wb.sheetnames[0]]
        print(f"📄 Using sheet: {ws.title}")
        print(f"📊 Total rows: {ws.max_row:,}")
        
        # Read header row
        headers = []
        header_row = ws[1]
        for cell in header_row:
            if cell.value:
                headers.append(str(cell.value).strip())
            else:
                break
        
        print(f"📋 Found {len(headers)} columns: {headers[:10]}...")
        
        # Map column names to indices
        col_map = {}
        for idx, header in enumerate(headers, start=1):
            col_map[header.lower()] = idx
        
        # Required columns
        required_cols = ['ingredient_id', 'region_code', 'package_size_value_si', 'package_unit_si']
        missing_cols = [col for col in required_cols if col not in col_map]
        if missing_cols:
            print(f"⚠️  WARNING: Missing columns: {missing_cols}")
            print(f"Available columns: {list(col_map.keys())}")
        
        # Read data rows and write incrementally to avoid memory issues
        total_rows = ws.max_row
        batch_size = 250  # Show progress every 250 rows (more frequent updates)
        output_file = os.path.join(OUTPUT_DIR, "grocery_item_catalog.json")
        
        print(f"🔄 Processing {total_rows - 1:,} data rows...")
        print(f"📝 Writing to: {output_file}")
        print(f"⏱️  Progress updates every {batch_size} rows\n")
        
        import time
        start_time = time.time()
        row_count = 0
        valid_items = 0
        
        # Open output file for incremental writing
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write('[\n')
            first_item = True
            
            for row_idx in range(2, total_rows + 1):
                # Show progress frequently (flush immediately)
                if row_count > 0 and row_count % batch_size == 0:
                    elapsed = time.time() - start_time
                    progress = (row_count / (total_rows - 1)) * 100
                    rate = row_count / elapsed if elapsed > 0 else 0
                    remaining = (total_rows - 1 - row_count) / rate if rate > 0 else 0
                    print(f"   ⏳ Row {row_idx:,}/{total_rows:,} | Valid: {valid_items:,} | Progress: {progress:.1f}% | Rate: {rate:.0f} rows/sec | ETA: {remaining:.0f}s", flush=True)
                
                row_count += 1
                
                # Get ingredient_id first (this is the critical field)
                ingredient_id = ws.cell(row_idx, col_map.get('ingredient_id', 2)).value
                if not ingredient_id:
                    continue  # Skip rows without ingredient_id
                
                # Get region_code (also required)
                region_code = ws.cell(row_idx, col_map.get('region_code', 5)).value
                if not region_code:
                    continue  # Skip rows without region_code
                
                # Extract all fields
                item = {
                    "grocery_item_id": ws.cell(row_idx, col_map.get('grocery_item_id', 1)).value or "",
                    "ingredient_id": str(ingredient_id).strip(),
                    "primary_name": ws.cell(row_idx, col_map.get('primary_name', 3)).value or "",
                    "category": ws.cell(row_idx, col_map.get('category', 4)).value or "",
                    "region_code": ws.cell(row_idx, col_map.get('region_code', 5)).value or "",
                    "package_size_value_SI": None,
                    "package_unit_SI": ws.cell(row_idx, col_map.get('package_unit_si', 7)).value or "",
                    "package_label_display": ws.cell(row_idx, col_map.get('package_label_display', 8)).value or "",
                    "priority_rank": None,
                    "typical_use": ws.cell(row_idx, col_map.get('typical_use', 10)).value or "",
                    "buy_as": ws.cell(row_idx, col_map.get('buy_as', 11)).value or "",
                    "notes_internal": ws.cell(row_idx, col_map.get('notes_internal', 12)).value or "",
                    "piece_size_SI_value": None,
                }
                
                # Handle numeric fields
                package_size_val = ws.cell(row_idx, col_map.get('package_size_value_si', 6)).value
                if package_size_val is not None:
                    try:
                        item["package_size_value_SI"] = float(package_size_val)
                    except (ValueError, TypeError):
                        pass
                
                priority_rank_val = ws.cell(row_idx, col_map.get('priority_rank', 9)).value
                if priority_rank_val is not None:
                    try:
                        item["priority_rank"] = int(priority_rank_val)
                    except (ValueError, TypeError):
                        pass
                
                piece_size_val = ws.cell(row_idx, col_map.get('piece_size_si_value', 13)).value
                if piece_size_val is not None:
                    try:
                        item["piece_size_SI_value"] = float(piece_size_val)
                    except (ValueError, TypeError):
                        pass
                
                # Skip if missing critical fields
                if not item["ingredient_id"] or not item["region_code"]:
                    continue
                
                # Write item immediately (incremental)
                if not first_item:
                    f.write(',\n')
                json.dump(item, f, ensure_ascii=False)
                first_item = False
                valid_items += 1
            
            f.write('\n]')
        
        elapsed_total = time.time() - start_time
        print(f"\n✅ Completed!")
        print(f"   📊 Processed {row_count:,} rows")
        print(f"   ✅ Valid items: {valid_items:,}")
        print(f"   ⏱️  Total time: {elapsed_total:.1f} seconds ({elapsed_total/60:.1f} minutes)")
        print(f"   📝 Output: {output_file}")
        
        # Calculate file size and statistics
        file_size = os.path.getsize(output_file)
        print(f"   📦 File size: {file_size / 1024 / 1024:.2f} MB ({file_size / 1024:.1f} KB)")
        
        # Read back a sample to get statistics (without loading everything)
        print(f"\n📊 Calculating statistics...")
        regions = set()
        ingredients = set()
        sample_count = 0
        
        with open(output_file, 'r', encoding='utf-8') as f:
            import json as json_lib
            data = json_lib.load(f)
            for item in data[:1000]:  # Sample first 1000 for stats
                regions.add(item.get('region_code', ''))
                ingredients.add(item.get('ingredient_id', ''))
        
        # Get full stats by reading all (but we already have valid_items count)
        print(f"   - Regions: {len(regions)} ({', '.join(sorted(regions))})")
        print(f"   - Sample ingredients: {len(ingredients)}")
        
        return True
        
    except Exception as e:
        print(f"⚠️  ERROR reading Excel file: {e}")
        print("   Trying CSV fallback...")
        return False


def export_from_csv(filepath: str) -> bool:
    """Export from CSV file"""
    import csv
    
    try:
        grocery_items = []
        
        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            
            for row_idx, row in enumerate(reader, start=2):
                ingredient_id = row.get('ingredient_id', '').strip()
                if not ingredient_id:
                    continue
                
                # Extract all fields
                item = {
                    "grocery_item_id": row.get('grocery_item_id', '').strip(),
                    "ingredient_id": ingredient_id,
                    "primary_name": row.get('primary_name', '').strip(),
                    "category": row.get('category', '').strip(),
                    "region_code": row.get('region_code', '').strip(),
                    "package_size_value_SI": None,
                    "package_unit_SI": row.get('package_unit_SI', '').strip(),
                    "package_label_display": row.get('package_label_display', '').strip(),
                    "priority_rank": None,
                    "typical_use": row.get('typical_use', '').strip(),
                    "buy_as": row.get('buy_as', '').strip(),
                    "notes_internal": row.get('notes_internal', '').strip(),
                    "piece_size_SI_value": None,
                }
                
                # Handle numeric fields
                try:
                    if row.get('package_size_value_SI'):
                        item["package_size_value_SI"] = float(row['package_size_value_SI'])
                except (ValueError, TypeError):
                    pass
                
                try:
                    if row.get('priority_rank'):
                        item["priority_rank"] = int(row['priority_rank'])
                except (ValueError, TypeError):
                    pass
                
                try:
                    if row.get('piece_size_SI_value'):
                        item["piece_size_SI_value"] = float(row['piece_size_SI_value'])
                except (ValueError, TypeError):
                    pass
                
                # Skip if missing critical fields
                if not item["ingredient_id"] or not item["region_code"]:
                    continue
                
                grocery_items.append(item)
                
                if len(grocery_items) % 100 == 0:
                    print(f"  Processed {len(grocery_items)} rows...")
        
        print(f"✅ Read {len(grocery_items)} grocery items from CSV")
        
        # Write to JSON
        output_file = os.path.join(OUTPUT_DIR, "grocery_item_catalog.json")
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(grocery_items, f, indent=2, ensure_ascii=False)
        
        print(f"✅ Exported {len(grocery_items)} grocery items")
        print(f"📝 Output: {output_file}")
        
        # Calculate file size
        file_size = os.path.getsize(output_file)
        print(f"📦 Size: {file_size / 1024:.1f} KB")
        
        # Print sample statistics
        regions = set(item["region_code"] for item in grocery_items)
        ingredients = set(item["ingredient_id"] for item in grocery_items)
        print(f"📊 Statistics:")
        print(f"   - Regions: {len(regions)} ({', '.join(sorted(regions))})")
        print(f"   - Unique ingredients: {len(ingredients)}")
        
        return True
        
    except Exception as e:
        print(f"❌ ERROR reading CSV file: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """Main execution"""
    print("=" * 80)
    print("GROCERY ITEM CATALOG EXPORTER")
    print("=" * 80)
    
    create_output_dir()
    export_grocery_catalog()
    
    print("\n" + "=" * 80)
    print("✅ EXPORT COMPLETE")
    print("=" * 80)


if __name__ == "__main__":
    main()

